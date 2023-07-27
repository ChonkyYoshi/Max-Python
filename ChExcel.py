import openpyxl as xl
from openpyxl.utils.cell import get_column_letter
import regex as reg
import bs4
from datetime import date


def GetDefinedNameData(defined_name):
    Name = dict()
    Name['Name'] = defined_name.name
    Name['Type'] = defined_name.type
    Name['Value'] = defined_name.value
    return Name


def GetCommentData(Comment):
    Com = dict()
    Com['Content'] = Comment.content
    Com['Author'] = Comment.author
    return Com


def GetHyperlinkData(cell):
    Link = dict()
    Link['ref'] = cell.hyperlink.ref
    Link['Location'] = cell.hyperlink.location
    Link['Tooltip'] = cell.hyperlink.tooltip
    Link['display'] = cell.hyperlink.display
    Link['Target'] = cell.hyperlink.target
    Link['Text'] = cell.internal_value
    return Link


def GetHiddenRows(Sheet):
    hidden_rows = list()
    for row in range(Sheet.min_row, Sheet.max_row + 1):
        if Sheet.row_dimensions[row].hidden:
            if row not in hidden_rows:
                hidden_rows.append(row)
    return hidden_rows


def GetHiddenColumns(Sheet):
    hidden_columns = list()
    for col in range(Sheet.min_column, Sheet.max_column + 1):
        if Sheet.column_dimensions[get_column_letter(col)].hidden:
            if col not in hidden_columns:
                hidden_columns.append(col)
                if Sheet.column_dimensions[get_column_letter(col)]\
                        .outlineLevel == 1:
                    for i in range(Sheet.column_dimensions[get_column_letter
                                                           (col)].min,
                                   Sheet.column_dimensions[get_column_letter
                                                           (col)].max):
                        if i not in hidden_columns:
                            hidden_columns.append(i)
    return hidden_columns


def GetTableData(table):
    Table = dict()
    Table['Name'] = table.name
    Table['Display Name'] = table.displayName
    Table['Range'] = table.ref
    return Table


def GetDataValidationData(datavalidation):
    Infos = dict()
    Infos['Prompt'] = datavalidation.prompt
    Infos['Prompt Title'] = datavalidation.promptTitle
    Infos['Error message'] = datavalidation.error
    Infos['Error Style'] = datavalidation.errorStyle
    Infos['Error Title'] = datavalidation.errorTitle
    Infos['Formula'] = datavalidation.formula1
    Infos['Type'] = datavalidation.type
    Infos['Range'] = list()
    for range in datavalidation.cells.ranges:
        Infos['Range'].append(range.coord)
    return Infos


Results = dict()
wb = xl.open('C:/Users/transformers/Downloads/FlagsAll.xlsm', keep_vba=True)
Results['Global defined names'] = list()
for name in wb.defined_names:
    Results['Global defined names'].append(
        GetDefinedNameData(wb.defined_names[name]))
for index, ws in enumerate(wb.worksheets):
    Results[ws.title] = dict(Visibility=ws.sheet_state,
                             Protection=ws.protection.sheet,
                             HiddenRows=GetHiddenRows(ws),
                             HiddenColumns=GetHiddenColumns(ws),
                             DefinedNames=list(), Tables=list(),
                             Comments=list(), Hyperlinks=list(),
                             Formulas=list(), HTML=list(),
                             Placeholder=list(), Long=list())
    if len(ws.defined_names) > 0:
        Results[ws.title]['Sheet defined names'] = list()
        for name in ws.defined_names:
            Results[ws.title]['DefinedNames'].append(
                GetDefinedNameData(ws.defined_names[name]))
    if len(ws.tables) > 0:
        Results[ws.title]['Sheet Tables'] = list()
        for table in ws.tables:
            Results[ws.title]['Sheet Tables'].append(
                GetTableData(ws.tables[table]))
    if len(ws.data_validations.dataValidation) > 0:
        Results[ws.title]['Data Validation'] = list()
        for data in ws.data_validations.dataValidation:
            Results[ws.title]['Data Validation'].append(
                GetDataValidationData(data))
    for row in range(ws.min_row, ws.max_row + 1):
        for col in range(ws.min_column, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.comment is not None:
                Results[ws.title]['Comments'].append(dict(
                    Cell=cell.coordinate,
                    CommentData=GetCommentData(cell.comment)))
            if cell.data_type == 'f':
                if cell.internal_value.startswith('='):
                    Results[ws.title]['Formulas'].append(
                        dict(Cell=cell.coordinate,
                             Formula=str(cell.internal_value)))
            if reg.search(r'<[^>]+>', str(cell.internal_value)):
                Results[ws.title]['HTML'].append(cell.coordinate)
            if reg.search(r'{.*?}', str(cell.internal_value)):
                Results[ws.title]['Placeholder'].append(cell.coordinate)
            if cell.hyperlink != '' and cell.hyperlink is not None:
                Results[ws.title]['Hyperlinks'].append(
                    dict(Cell=cell.coordinate,
                         HyperlinkData=GetHyperlinkData(cell)))
            if len(str(cell.internal_value)) > 25000:
                Results[ws.title]['Long'].append(cell.coordinate)

with open('base.html') as File:
    Report = bs4.BeautifulSoup(File, 'html.parser')
Report.head.title.contents = f'Excel Preflight Report {date.today}'
h1 = Report.new_tag('H1')
h1.contents = 'Excel Preflight Report'
Report.body.append()
with open('Report.html', 'x') as File:
    File.write(bs4.BeautifulSoup.prettify(Report))
1 == 1
